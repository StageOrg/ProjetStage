# === IMPORT MASSIF ANCIENS ÉTUDIANTS (CSV, EXCEL, PDF) ===
import csv
from io import StringIO, BytesIO
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

from apps.inscription_pedagogique.models import (
    AnneeAcademique, Inscription, Parcours, Filiere, AnneeEtude
)
from apps.utilisateurs.models import Etudiant
from apps.page_professeur.models import UE, ResultatUE

# === EXCEL & PDF LIBS ===
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_anciens_etudiants(request):
    fichier = request.FILES.get('fichier')
    parcours_id = request.data.get('parcours_id')
    filiere_id = request.data.get('filiere_id')
    annee_etude_id = request.data.get('annee_etude_id')
    format_export = request.data.get('format', 'csv').lower()

    # === VALIDATION ===
    if not all([fichier, parcours_id, filiere_id, annee_etude_id]):
        return Response(
            {'error': 'Tous les champs sont requis : fichier, parcours, filière, année'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        parcours = Parcours.objects.get(id=parcours_id)
        filiere = Filiere.objects.get(id=filiere_id)
        annee_etude = AnneeEtude.objects.get(id=annee_etude_id)
        annee_active = AnneeAcademique.objects.get(est_active=True)
    except Exception:
        return Response({'error': 'Filtres invalides'}, status=status.HTTP_400_BAD_REQUEST)

    # === LIRE LE FICHIER (CSV, XLSX, PDF) ===
    try:
        file_ext = fichier.name.split('.')[-1].lower()
        if file_ext == 'csv':
            content = fichier.read().decode('utf-8-sig')
            csv_reader = csv.DictReader(StringIO(content))
            rows = list(csv_reader)
        elif file_ext in ['xlsx', 'xls']:
            import pandas as pd
            df = pd.read_excel(fichier)
            rows = df.to_dict('records')
        elif file_ext == 'pdf':
            import PyPDF2
            reader = PyPDF2.PdfReader(fichier)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            # Simple parsing: assume one num_carte per line
            rows = [{'num_carte': line.strip()} for line in text.split('\n') if line.strip().isdigit() and len(line.strip()) == 6]
        else:
            return Response({'error': 'Format non supporté'}, status=400)
    except Exception as e:
        return Response({'error': f'Fichier invalide : {str(e)}'}, status=400)

    resultats = []
    reussis = 0
    echoues = 0

    # === FONCTION : UES DISPONIBLES (≤ 70 CRÉDITS) ===
    def get_ues_disponibles(etudiant):
        derniere_inscription = Inscription.objects.filter(
            etudiant=etudiant
        ).select_related('parcours', 'filiere', 'annee_etude').order_by('-anneeAcademique__libelle').first()

        if not derniere_inscription:
            return []

        ues_non_validees = ResultatUE.objects.filter(
            etudiant=etudiant,
            inscription=derniere_inscription,
            est_valide=False
        ).select_related('ue')

        credits_rattrapage = sum(ue.ue.nbre_credit for ue in ues_non_validees)
        ues_selectionnees = [
            {'id': ue.ue.id, 'credits': ue.ue.nbre_credit, 'type': 'rattrapage'}
            for ue in ues_non_validees
        ]

        if credits_rattrapage > 70:
            return ues_selectionnees

        credits_restants = 70 - credits_rattrapage
        ues_nouvelle_annee = UE.objects.filter(
            parcours=parcours,
            filiere=filiere,
            annee_etude=annee_etude
        ).exclude(id__in=[u['id'] for u in ues_selectionnees])

        for ue in ues_nouvelle_annee.order_by('code'):
            if credits_restants - ue.nbre_credit >= 0:
                ues_selectionnees.append({
                    'id': ue.id,
                    'credits': ue.nbre_credit,
                    'type': 'nouvelle_annee'
                })
                credits_restants -= ue.nbre_credit
            else:
                break
        return ues_selectionnees

    # === TRAITEMENT ===
    for row in rows:
        num_carte = str(row.get('num_carte', '')).strip()
        if not num_carte or len(num_carte) != 6 or not num_carte.isdigit():
            resultats.append({'num_carte': num_carte, 'statut': 'échoué', 'erreur': 'Numéro invalide'})
            echoues += 1
            continue

        try:
            etudiant = Etudiant.objects.get(num_carte=num_carte)
            ues_disponibles = get_ues_disponibles(etudiant)

            if not ues_disponibles:
                resultats.append({'num_carte': num_carte, 'statut': 'échoué', 'erreur': 'Aucune UE'})
                echoues += 1
                continue

            total_credits = sum(ue['credits'] for ue in ues_disponibles)
            inscription_existante = Inscription.objects.filter(
                etudiant=etudiant, anneeAcademique=annee_active
            ).first()

            if inscription_existante:
                resultats.append({
                    'num_carte': num_carte,
                    'statut': 'échoué',
                    'erreur': f'Déjà inscrit ({inscription_existante.numero})'
                })
                echoues += 1
                continue

            numero_inscription = f"INS-{num_carte}-{annee_active.libelle}"
            inscription = Inscription.objects.create(
                numero=numero_inscription,
                etudiant=etudiant,
                parcours=parcours,
                filiere=filiere,
                annee_etude=annee_etude,
                anneeAcademique=annee_active
            )

            for ue in ues_disponibles:
                inscription.ues.add(UE.objects.get(id=ue['id']))

            resultats.append({
                'num_carte': num_carte,
                'statut': 'réussi',
                'nom': f"{etudiant.utilisateur.last_name} {etudiant.utilisateur.first_name}",
                'ues_inscrites': len(ues_disponibles),
                'total_credits': total_credits,
                'rattrapage': sum(1 for u in ues_disponibles if u['type'] == 'rattrapage'),
                'nouvelle_annee': sum(1 for u in ues_disponibles if u['type'] == 'nouvelle_annee'),
                'numero_inscription': numero_inscription
            })
            reussis += 1

        except Etudiant.DoesNotExist:
            resultats.append({'num_carte': num_carte, 'statut': 'échoué', 'erreur': 'Non trouvé'})
            echoues += 1
        except Exception as e:
            resultats.append({'num_carte': num_carte, 'statut': 'échoué', 'erreur': str(e)})
            echoues += 1

    # === GÉNÉRER LE FICHIER DE RÉSULTAT ===
    if format_export == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultat Import"
        headers = ['N° Carte', 'Statut', 'Nom', 'UEs', 'Crédits', 'Rattr.', 'Nouv.', 'N° Insc.', 'Erreur']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for r in resultats:
            ws.append([
                r.get('num_carte', ''),
                r.get('statut', ''),
                r.get('nom', ''),
                r.get('ues_inscrites', ''),
                r.get('total_credits', ''),
                r.get('rattrapage', ''),
                r.get('nouvelle_annee', ''),
                r.get('numero_inscription', ''),
                r.get('erreur', '')
            ])

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                if cell.column == 2:
                    cell.font = Font(color="00AA00" if cell.value == "réussi" else "E53935")

        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.getvalue()
        filename = "resultat_import_anciens.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    elif format_export == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Résultat Import Anciens Étudiants", styles['Title']))
        elements.append(Spacer(1, 12))

        data = [['N° Carte', 'Statut', 'Nom', 'UEs', 'Crédits', 'Rattr.', 'Nouv.', 'N° Insc.', 'Erreur']]
        for r in resultats:
            data.append([r.get(k, '') for k in ['num_carte', 'statut', 'nom', 'ues_inscrites', 'total_credits', 'rattrapage', 'nouvelle_annee', 'numero_inscription', 'erreur']])

        table = Table(data, colWidths=[60, 50, 100, 40, 50, 40, 40, 80, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E88E5")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,1), (-1,-1), 9),
        ]))
        for i, r in enumerate(resultats, 1):
            if r.get('statut') == 'réussi':
                table.setStyle(TableStyle([('TEXTCOLOR', (1,i), (1,i), colors.green)]))
            else:
                table.setStyle(TableStyle([('TEXTCOLOR', (1,i), (1,i), colors.red)]))
        elements.append(table)
        doc.build(elements)
        content = buffer.getvalue()
        filename = "resultat_import_anciens.pdf"
        content_type = "application/pdf"

    else:  # CSV
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['num_carte', 'statut', 'nom', 'ues_inscrites', 'total_credits', 'rattrapage', 'nouvelle_annee', 'numero_inscription', 'erreur'])
        for r in resultats:
            writer.writerow([r.get(k, '') for k in ['num_carte', 'statut', 'nom', 'ues_inscrites', 'total_credits', 'rattrapage', 'nouvelle_annee', 'numero_inscription', 'erreur']])
        content = output.getvalue()
        filename = "resultat_import_anciens.csv"
        content_type = "text/csv"

    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response